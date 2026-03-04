"""Parsing utilities for questionnaires and reference documents."""

import re
import csv
import io


def parse_questionnaire(filepath: str, ext: str) -> list[dict]:
    """Parse a questionnaire file into a list of {text, location_meta} dicts."""
    if ext == "pdf":
        return _parse_questionnaire_pdf(filepath)
    elif ext == "xlsx":
        return _parse_questionnaire_xlsx(filepath)
    elif ext == "txt":
        return _parse_questionnaire_txt(filepath)
    return []


def _parse_questionnaire_pdf(filepath: str) -> list[dict]:
    import pdfplumber

    # Collect all lines with page info
    all_lines: list[tuple[str, int]] = []
    with pdfplumber.open(filepath) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            for line in text.split("\n"):
                stripped = line.strip()
                if stripped:
                    all_lines.append((stripped, page_num))

    return _merge_and_extract_questions(all_lines)


def _parse_questionnaire_xlsx(filepath: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    all_lines: list[tuple[str, int]] = []
    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell and isinstance(cell, str):
                stripped = cell.strip()
                if stripped:
                    all_lines.append((stripped, row_num))
    wb.close()
    return _merge_and_extract_questions(all_lines)


def _parse_questionnaire_txt(filepath: str) -> list[dict]:
    all_lines: list[tuple[str, int]] = []
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        for line_num, line in enumerate(f, start=1):
            stripped = line.strip()
            if stripped:
                all_lines.append((stripped, line_num))
    return _merge_and_extract_questions(all_lines)


# ---- question numbering pattern ----
# Matches: "Q1.", "Q1)", "Q1:", "1.", "1)", "1:", etc. at the start of a line
_QUESTION_START_RE = re.compile(r"^Q?\d+[\.\)\:]\s", re.IGNORECASE)


def _starts_new_question(text: str) -> bool:
    """Check if a line begins a new numbered question."""
    return bool(_QUESTION_START_RE.match(text))


def _merge_and_extract_questions(lines: list[tuple[str, int]]) -> list[dict]:
    """Merge multi-line questions and extract them.

    Strategy:
      1. Lines that start with a question number (Q1., 1., 2), etc.)
         begin a new question.
      2. Subsequent lines that do NOT start with a question number are
         treated as continuation of the previous question.
      3. After merging, any non-question preamble text is discarded.
    """
    # Phase 1: group lines into blocks (new-question-start → continuation lines)
    blocks: list[dict] = []  # {"lines": [str], "location": str/int}
    for text, loc in lines:
        if _starts_new_question(text):
            # Begin a new block
            blocks.append({"lines": [text], "location": loc})
        elif blocks:
            # Continuation of the current block
            blocks[-1]["lines"].append(text)
        # else: preamble text before any question — skip

    # Phase 2: merge lines within each block and clean
    questions = []
    for block in blocks:
        merged = " ".join(block["lines"])
        cleaned = _clean_question(merged)
        if cleaned and len(cleaned) >= 10:
            loc = block["location"]
            loc_label = f"page {loc}" if isinstance(loc, int) else str(loc)
            questions.append({"text": cleaned, "location_meta": loc_label})

    # Fallback: if no numbered questions were found, try the old
    # line-by-line heuristic (handles free-form question lists ending with ?)
    if not questions:
        for text, loc in lines:
            if _looks_like_question_fallback(text):
                cleaned = _clean_question(text)
                if cleaned and len(cleaned) >= 10:
                    loc_label = f"page {loc}" if isinstance(loc, int) else str(loc)
                    questions.append({"text": cleaned, "location_meta": loc_label})

    return questions


def _looks_like_question_fallback(text: str) -> bool:
    """Fallback heuristic for files without numbered questions.

    Only triggers when _merge_and_extract_questions found zero numbered
    questions.  Uses stricter rules to avoid false positives on fragments.
    """
    if not text or len(text) < 20:
        return False
    # Must end with ? and be long enough to be a real question
    if text.endswith("?") and len(text) >= 20:
        return True
    return False


def _clean_question(text: str) -> str:
    """Remove leading numbering but keep the question text."""
    cleaned = re.sub(r"^Q?\d+[\.\)\:]\s*", "", text, flags=re.IGNORECASE).strip()
    return cleaned if cleaned else text


# ---- reference text extraction ----

def extract_reference_text(filepath: str, ext: str) -> str:
    if ext == "pdf":
        return _extract_pdf(filepath)
    elif ext == "txt":
        return _extract_txt(filepath)
    elif ext == "csv":
        return _extract_csv(filepath)
    elif ext == "docx":
        return _extract_docx(filepath)
    return ""


def _extract_pdf(filepath: str) -> str:
    import pdfplumber

    texts = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            texts.append(t)
    # Use form-feed between pages so the splitter can detect page boundaries
    return "\f".join(texts)


def _extract_txt(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _extract_csv(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        rows = [", ".join(row) for row in reader]
    return "\n".join(rows)


def _extract_docx(filepath: str) -> str:
    from docx import Document

    doc = Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
